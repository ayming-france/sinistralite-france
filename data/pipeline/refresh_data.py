#!/usr/bin/env python3
"""Telecharge les donnees AT + MP depuis ameli.fr et produit les JSON pour le dashboard.

Usage:
    python refresh_data.py [--pdf-dir /chemin/vers/pdfs]

Sorties dans data/ (dossier parent de data/pipeline/) :
    at-data.json, mp-data.json, trajet-data.json
"""

import json
import subprocess
import sys
from pathlib import Path
from collections import defaultdict

PIPELINE_DIR = Path(__file__).parent
OUTPUT_DIR = PIPELINE_DIR.parent  # data/

# ── AT config ──
AT_XLSX_PATH = PIPELINE_DIR / "at-by-ctn-naf.xlsx"
AT_JSON_PATH = OUTPUT_DIR / "at-data.json"
AT_XLSX_URL = (
    "https://assurance-maladie.ameli.fr/sites/default/files/2023_Risque-AT-CTN-x-NAF_serie%20annuelle.xlsx"
)
# Le fichier 2021 est necessaire pour les colonnes de causes de risque (colonnes 20-31)
# qui ne sont pas disponibles dans le fichier 2023 au meme format.
AT_2021_XLSX_PATH = PIPELINE_DIR / "at-2021.xlsx"
AT_2021_XLSX_URL = (
    "https://www.assurance-maladie.ameli.fr/sites/default/files/2021_Risque-AT-CTN-x-NAF_serie%20annuelle.xlsx"
)

AT_COL = {
    "ctn": 0, "libelle_ctn": 1,
    "naf5": 2, "libelle_naf": 3,
    "naf2": 4, "libelle_naf2": 5,
    "naf38": 6, "libelle_naf38": 7,
    "nb_salaries": 8, "nb_heures": 9, "nb_siret": 10,
    "at_1er_reglement": 11, "at_4j_arret": 12,
    "nouvelles_ip": 13, "deces": 16, "journees_it": 17,
}
AT_RISK_CAUSES = {
    20: "Manutention manuelle", 21: "Chutes de plain-pied",
    22: "Risque chimique", 23: "Chutes de hauteur",
    24: "Risque physique", 25: "Risque machines",
    26: "Outillage a main", 27: "Risque routier",
    28: "Agressions", 29: "Manutention mecanique",
    30: "Autres risques", 31: "Autres vehicules",
}

# ── MP config ──
MP_XLSX_PATH = PIPELINE_DIR / "mp-by-ctn-naf.xlsx"
MP_JSON_PATH = OUTPUT_DIR / "mp-data.json"
MP_XLSX_URL = (
    "https://assurance-maladie.ameli.fr/sites/default/files/2023_Risque-MP-par-CTN-x-NAF_serie-annuelle.xlsx"
)

# MP Excel: NAF38 at 4-5, NAF2 at 6-7 (swapped vs AT)
MP_COL = {
    "ctn": 0, "libelle_ctn": 1,
    "naf5": 2, "libelle_naf": 3,
    "naf38": 4, "libelle_naf38": 5,
    "naf2": 6, "libelle_naf2": 7,
    "tableau_code": 8, "tableau_libelle": 9,
    "syndrome_code": 10, "syndrome_libelle": 11,
    "nb_salaries": 12, "nb_heures": 13, "nb_siret": 14,
    "mp_1er_reglement": 15, "nouvelles_ip": 16,
    "ip_taux_inf_10": 17, "ip_taux_sup_10": 18,
    "deces": 19, "journees_it": 20,
    "somme_taux_ip": 21,
    "top_tms": 24, "top_biologie": 25, "top_chimique": 26,
    "top_cancers": 27, "top_cancers_ht": 28, "top_psy": 29,
}

# Causes MP derivees des flags (colonnes 24-29)
MP_CAUSE_FLAGS = {
    "TMS": 24,
    "Risque chimique": 26,
    "Cancers professionnels": 27,
    "Risque biologique": 25,
    "Risque psychosocial": 29,
}
# Fallback pour les lignes sans flag
MP_TABLEAU_MAP = {
    "Hors tableau": "Hors tableau",
    "Atteinte auditive": "Bruit",
    "Affections oculaire": "Autres MP",
    "Affections provoquees par les rayonnements": "Autres MP",
    "Lesions provoquees": "Autres MP",
}

# ── Trajet config ──
TRAJET_JSON_PATH = OUTPUT_DIR / "trajet-data.json"

HEADER_ROW = 4  # 1-based, identique pour AT et MP


def download_xlsx(path, url):
    """Telecharge un fichier Excel depuis ameli.fr via curl (gere le SSL)."""
    if path.exists():
        print(f"  [skip] Deja present : {path.name}")
        return
    print(f"  [dl] Telechargement : {path.name} ...")
    subprocess.run(["curl", "-sL", "-o", str(path), url], check=True)
    print(f"  [ok] Sauvegarde ({path.stat().st_size / 1024:.0f} KB)")


def safe_num(val):
    if val is None:
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


def write_json(data, json_path, label):
    """Ecrit un fichier JSON (seul format de sortie, sans pickle)."""
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  [ok] {label} : {json_path.name} ({json_path.stat().st_size / 1024:.0f} KB)")


# ═══════════════════════════════════════════
# AT PIPELINE
# ═══════════════════════════════════════════

def parse_at_xlsx():
    """Parse le fichier Excel AT, retourne une liste de dicts par ligne."""
    import openpyxl
    wb = openpyxl.load_workbook(AT_XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active

    data_rows = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        naf5_val = row[AT_COL["naf5"]]
        if not naf5_val:
            continue
        naf5 = str(naf5_val).strip()
        if not naf5 or naf5 == "None":
            continue

        entry = {
            "ctn": str(row[AT_COL["ctn"]] or "").strip(),
            "naf5": naf5,
            "naf2": str(row[AT_COL["naf2"]] or naf5[:2]).strip(),
            "libelle": str(row[AT_COL["libelle_naf"]] or "").strip(),
            "libelle_naf2": str(row[AT_COL["libelle_naf2"]] or "").strip(),
        }

        for key in ["nb_salaries", "nb_heures", "nb_siret", "at_1er_reglement",
                     "at_4j_arret", "nouvelles_ip", "deces", "journees_it"]:
            entry[key] = safe_num(row[AT_COL[key]])

        entry["risk_causes_raw"] = {}
        for col_idx, label in AT_RISK_CAUSES.items():
            entry["risk_causes_raw"][label] = safe_num(row[col_idx])

        data_rows.append(entry)

    wb.close()
    print(f"  [parse] {len(data_rows)} lignes AT")
    return data_rows


def compute_at_stats(group):
    """Calcule les statistiques AT derivees depuis les sommes brutes."""
    nb_sal = group["nb_salaries"]
    nb_h = group["nb_heures"]
    at_4j = group["at_4j_arret"]
    return {
        "nb_salaries": int(nb_sal),
        "nb_heures": int(nb_h),
        "nb_siret": int(group["nb_siret"]),
        "at_1er_reglement": int(group["at_1er_reglement"]),
        "at_4j_arret": int(at_4j),
        "nouvelles_ip": int(group["nouvelles_ip"]),
        "deces": int(group["deces"]),
        "journees_it": int(group["journees_it"]),
        "indice_frequence": round(at_4j / nb_sal * 1000, 1) if nb_sal > 0 else 0,
        "taux_gravite": round(group["journees_it"] / (nb_h / 1000), 2) if nb_h > 0 else 0,
    }


def compute_risk_causes(causes_summed, total_at_4j):
    """Convertit les comptages bruts de causes AT en pourcentages."""
    if total_at_4j == 0:
        return {name: 0 for name in causes_summed}
    return {name: round(count / total_at_4j * 100, 1) for name, count in causes_summed.items()}


def aggregate_at_rows(rows, key_fn, libelle_fn=None):
    """Agrege les lignes AT par une cle de groupement."""
    groups = defaultdict(lambda: {
        "nb_salaries": 0, "nb_heures": 0, "nb_siret": 0,
        "at_1er_reglement": 0, "at_4j_arret": 0,
        "nouvelles_ip": 0, "deces": 0, "journees_it": 0,
        "risk_causes_raw": defaultdict(float),
        "source_codes": [],
        "libelle": "",
    })
    for row in rows:
        key = key_fn(row)
        if not key:
            continue
        g = groups[key]
        for field in ["nb_salaries", "nb_heures", "nb_siret", "at_1er_reglement",
                       "at_4j_arret", "nouvelles_ip", "deces", "journees_it"]:
            g[field] += row[field]
        for cause, val in row["risk_causes_raw"].items():
            g["risk_causes_raw"][cause] += val
        g["source_codes"].append(row["naf5"])
        if libelle_fn:
            g["libelle"] = libelle_fn(row, g["libelle"])
        elif not g["libelle"]:
            g["libelle"] = row.get("libelle", "")
    return groups


def build_at_data(rows):
    """Construit la structure de donnees AT (NAF5, NAF4, NAF2 + index)."""
    naf5_groups = aggregate_at_rows(rows, key_fn=lambda r: r["naf5"],
                                    libelle_fn=lambda r, prev: prev or r["libelle"])
    by_naf5 = {}
    for code, g in sorted(naf5_groups.items()):
        by_naf5[code] = {
            "libelle": g["libelle"], "naf4": code[:4], "naf2": code[:2],
            "stats": compute_at_stats(g),
            "risk_causes": compute_risk_causes(dict(g["risk_causes_raw"]), g["at_4j_arret"]),
        }

    naf4_groups = aggregate_at_rows(rows, key_fn=lambda r: r["naf5"][:4],
                                    libelle_fn=lambda r, prev: prev or r["libelle"])
    by_naf4 = {}
    for code, g in sorted(naf4_groups.items()):
        by_naf4[code] = {
            "libelle": g["libelle"], "naf2": code[:2],
            "codes_naf5": sorted(set(g["source_codes"])),
            "stats": compute_at_stats(g),
            "risk_causes": compute_risk_causes(dict(g["risk_causes_raw"]), g["at_4j_arret"]),
        }

    naf2_groups = aggregate_at_rows(rows, key_fn=lambda r: r["naf2"],
                                    libelle_fn=lambda r, prev: prev or r.get("libelle_naf2", r["libelle"]))
    by_naf2 = {}
    for code, g in sorted(naf2_groups.items()):
        by_naf2[code] = {
            "libelle": g["libelle"],
            "stats": compute_at_stats(g),
            "risk_causes": compute_risk_causes(dict(g["risk_causes_raw"]), g["at_4j_arret"]),
        }

    national = {k: 0 for k in ["nb_salaries", "nb_heures", "nb_siret",
                                 "at_1er_reglement", "at_4j_arret",
                                 "nouvelles_ip", "deces", "journees_it"]}
    for g in naf5_groups.values():
        for field in national:
            national[field] += g[field]

    naf_index = []
    for code, data in by_naf5.items():
        naf_index.append({"code": code, "libelle": data["libelle"], "level": "naf5"})
    for code, data in by_naf4.items():
        naf_index.append({"code": code, "libelle": data["libelle"], "level": "naf4"})
    for code, data in by_naf2.items():
        naf_index.append({"code": code, "libelle": data["libelle"], "level": "naf2"})
    naf_index.sort(key=lambda x: (x["level"], x["code"]))

    return {
        "meta": {
            "source": "Ameli - Risque AT par CTN x NAF 2023",
            "source_url": AT_XLSX_URL,
            "national": compute_at_stats(national),
        },
        "by_naf5": by_naf5, "by_naf4": by_naf4, "by_naf2": by_naf2,
        "naf_index": naf_index,
    }


# ═══════════════════════════════════════════
# MP PIPELINE
# ═══════════════════════════════════════════

def parse_mp_xlsx():
    """Parse le fichier Excel MP.
    - Main-d'oeuvre (nb_salaries, nb_heures, nb_siret) dedupliquee par CTN+NAF
    - Statistiques MP cumulees sur toutes les lignes (tableau x syndrome)
    - Causes trackees par ligne via les flags
    """
    import openpyxl
    wb = openpyxl.load_workbook(MP_XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Passe 1 : main-d'oeuvre unique par CTN+NAF
    workforce = {}
    # Passe 2 : cumul MP par NAF5
    naf5_mp = defaultdict(lambda: {
        "mp_1er_reglement": 0, "nouvelles_ip": 0,
        "ip_taux_inf_10": 0, "ip_taux_sup_10": 0,
        "deces": 0, "journees_it": 0, "somme_taux_ip": 0,
        "cause_counts": defaultdict(float),
        "libelle": "", "naf2": "", "libelle_naf2": "", "ctn_set": set(),
    })

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        naf5_val = row[MP_COL["naf5"]]
        if not naf5_val:
            continue
        naf5 = str(naf5_val).strip()
        if not naf5 or naf5 == "None":
            continue

        ctn = str(row[MP_COL["ctn"]] or "").strip()
        key = (ctn, naf5)

        # Main-d'oeuvre : premiere occurrence par CTN+NAF (identique sur toutes les lignes)
        if key not in workforce:
            workforce[key] = {
                "nb_salaries": safe_num(row[MP_COL["nb_salaries"]]),
                "nb_heures": safe_num(row[MP_COL["nb_heures"]]),
                "nb_siret": safe_num(row[MP_COL["nb_siret"]]),
            }

        # Statistiques MP : cumul sur toutes les lignes (tableau x syndrome)
        mp = naf5_mp[naf5]
        mp_count = safe_num(row[MP_COL["mp_1er_reglement"]])
        mp["mp_1er_reglement"] += mp_count
        mp["nouvelles_ip"] += safe_num(row[MP_COL["nouvelles_ip"]])
        mp["ip_taux_inf_10"] += safe_num(row[MP_COL["ip_taux_inf_10"]])
        mp["ip_taux_sup_10"] += safe_num(row[MP_COL["ip_taux_sup_10"]])
        mp["deces"] += safe_num(row[MP_COL["deces"]])
        mp["journees_it"] += safe_num(row[MP_COL["journees_it"]])
        mp["somme_taux_ip"] += safe_num(row[MP_COL["somme_taux_ip"]])

        # Categorisation des causes : flags en priorite, puis nom du tableau
        if mp_count > 0:
            flagged = False
            for cause_name, col_idx in MP_CAUSE_FLAGS.items():
                if row[col_idx] == "oui":
                    mp["cause_counts"][cause_name] += mp_count
                    flagged = True
            if not flagged:
                tab_name = str(row[MP_COL["tableau_libelle"]] or "").strip()
                category = "Autres MP"
                for prefix, cat in MP_TABLEAU_MAP.items():
                    if tab_name.startswith(prefix):
                        category = cat
                        break
                mp["cause_counts"][category] += mp_count

        if not mp["libelle"]:
            mp["libelle"] = str(row[MP_COL["libelle_naf"]] or "").strip()
        if not mp["naf2"]:
            mp["naf2"] = str(row[MP_COL["naf2"]] or naf5[:2]).strip()
        if not mp["libelle_naf2"]:
            mp["libelle_naf2"] = str(row[MP_COL["libelle_naf2"]] or "").strip()
        mp["ctn_set"].add(ctn)

    wb.close()

    # Construction des lignes aggregees : une par NAF5, main-d'oeuvre somme sur les CTNs
    naf5_workforce = defaultdict(lambda: {"nb_salaries": 0, "nb_heures": 0, "nb_siret": 0})
    for (ctn, naf5), wf in workforce.items():
        for k in ["nb_salaries", "nb_heures", "nb_siret"]:
            naf5_workforce[naf5][k] += wf[k]

    data_rows = []
    for naf5, mp in naf5_mp.items():
        wf = naf5_workforce[naf5]
        data_rows.append({
            "naf5": naf5,
            "naf2": mp["naf2"],
            "libelle": mp["libelle"],
            "libelle_naf2": mp["libelle_naf2"],
            "nb_salaries": wf["nb_salaries"],
            "nb_heures": wf["nb_heures"],
            "nb_siret": wf["nb_siret"],
            "mp_1er_reglement": mp["mp_1er_reglement"],
            "nouvelles_ip": mp["nouvelles_ip"],
            "ip_taux_inf_10": mp["ip_taux_inf_10"],
            "ip_taux_sup_10": mp["ip_taux_sup_10"],
            "deces": mp["deces"],
            "journees_it": mp["journees_it"],
            "somme_taux_ip": mp["somme_taux_ip"],
            "cause_counts": dict(mp["cause_counts"]),
        })

    print(f"  [parse] {len(data_rows)} entrees MP NAF5 (depuis {sum(1 for _ in workforce)} combos CTN+NAF)")
    return data_rows


def compute_mp_stats(group):
    """Calcule les statistiques MP derivees."""
    nb_sal = group["nb_salaries"]
    nb_h = group["nb_heures"]
    mp = group["mp_1er_reglement"]
    return {
        "nb_salaries": int(nb_sal),
        "nb_heures": int(nb_h),
        "nb_siret": int(group["nb_siret"]),
        "mp_1er_reglement": int(mp),
        "nouvelles_ip": int(group["nouvelles_ip"]),
        "ip_taux_inf_10": int(group.get("ip_taux_inf_10", 0)),
        "ip_taux_sup_10": int(group.get("ip_taux_sup_10", 0)),
        "deces": int(group["deces"]),
        "journees_it": int(group["journees_it"]),
        "somme_taux_ip": int(group.get("somme_taux_ip", 0)),
        "indice_frequence": round(mp / nb_sal * 1000, 1) if nb_sal > 0 else 0,
        "taux_gravite": round(group["journees_it"] / (nb_h / 1000), 2) if nb_h > 0 else 0,
    }


def compute_mp_causes(cause_counts, total_mp):
    """Convertit les comptages de causes MP en pourcentages."""
    if total_mp == 0:
        return {name: 0 for name in cause_counts}
    return {name: round(count / total_mp * 100, 1) for name, count in cause_counts.items()}


def aggregate_mp_rows(rows, key_fn, libelle_fn=None):
    """Agrege les lignes MP par une cle de groupement."""
    groups = defaultdict(lambda: {
        "nb_salaries": 0, "nb_heures": 0, "nb_siret": 0,
        "mp_1er_reglement": 0, "nouvelles_ip": 0,
        "ip_taux_inf_10": 0, "ip_taux_sup_10": 0,
        "deces": 0, "journees_it": 0, "somme_taux_ip": 0,
        "cause_counts": defaultdict(float),
        "source_codes": [],
        "libelle": "",
    })
    for row in rows:
        key = key_fn(row)
        if not key:
            continue
        g = groups[key]
        for field in ["nb_salaries", "nb_heures", "nb_siret", "mp_1er_reglement",
                       "nouvelles_ip", "ip_taux_inf_10", "ip_taux_sup_10",
                       "deces", "journees_it", "somme_taux_ip"]:
            g[field] += row.get(field, 0)
        for cause, val in row.get("cause_counts", {}).items():
            g["cause_counts"][cause] += val
        g["source_codes"].append(row["naf5"])
        if libelle_fn:
            g["libelle"] = libelle_fn(row, g["libelle"])
        elif not g["libelle"]:
            g["libelle"] = row.get("libelle", "")
    return groups


def build_mp_data(rows):
    """Construit la structure de donnees MP (meme forme que AT pour le dashboard)."""
    naf5_groups = aggregate_mp_rows(rows, key_fn=lambda r: r["naf5"],
                                    libelle_fn=lambda r, prev: prev or r["libelle"])
    by_naf5 = {}
    for code, g in sorted(naf5_groups.items()):
        by_naf5[code] = {
            "libelle": g["libelle"], "naf4": code[:4], "naf2": code[:2],
            "stats": compute_mp_stats(g),
            "risk_causes": compute_mp_causes(dict(g["cause_counts"]), g["mp_1er_reglement"]),
        }

    naf4_groups = aggregate_mp_rows(rows, key_fn=lambda r: r["naf5"][:4],
                                    libelle_fn=lambda r, prev: prev or r["libelle"])
    by_naf4 = {}
    for code, g in sorted(naf4_groups.items()):
        by_naf4[code] = {
            "libelle": g["libelle"], "naf2": code[:2],
            "codes_naf5": sorted(set(g["source_codes"])),
            "stats": compute_mp_stats(g),
            "risk_causes": compute_mp_causes(dict(g["cause_counts"]), g["mp_1er_reglement"]),
        }

    naf2_groups = aggregate_mp_rows(rows, key_fn=lambda r: r["naf2"],
                                    libelle_fn=lambda r, prev: prev or r.get("libelle_naf2", r["libelle"]))
    by_naf2 = {}
    for code, g in sorted(naf2_groups.items()):
        by_naf2[code] = {
            "libelle": g["libelle"],
            "stats": compute_mp_stats(g),
            "risk_causes": compute_mp_causes(dict(g["cause_counts"]), g["mp_1er_reglement"]),
        }

    national = {k: 0 for k in ["nb_salaries", "nb_heures", "nb_siret",
                                 "mp_1er_reglement", "nouvelles_ip",
                                 "ip_taux_inf_10", "ip_taux_sup_10",
                                 "deces", "journees_it", "somme_taux_ip"]}
    for g in naf5_groups.values():
        for field in national:
            national[field] += g[field]

    naf_index = []
    for code, data in by_naf5.items():
        naf_index.append({"code": code, "libelle": data["libelle"], "level": "naf5"})
    for code, data in by_naf4.items():
        naf_index.append({"code": code, "libelle": data["libelle"], "level": "naf4"})
    for code, data in by_naf2.items():
        naf_index.append({"code": code, "libelle": data["libelle"], "level": "naf2"})
    naf_index.sort(key=lambda x: (x["level"], x["code"]))

    return {
        "meta": {
            "source": "Ameli - Risque MP par CTN x NAF 2023",
            "source_url": MP_XLSX_URL,
            "national": compute_mp_stats(national),
        },
        "by_naf5": by_naf5, "by_naf4": by_naf4, "by_naf2": by_naf2,
        "naf_index": naf_index,
    }


# ═══════════════════════════════════════════
# EVOLUTION ANNUELLE (parseurs legers)
# ═══════════════════════════════════════════

# Les colonnes de stats principales sont aux memes positions en 2021 et 2023 pour AT et MP.
# NAF5 est toujours en colonne 2. NAF2 derive de NAF5[:2].

YEARLY_AT_COLS = {
    "naf5": 2, "nb_salaries": 8, "nb_heures": 9, "nb_siret": 10,
    "events": 11,  # at_1er_reglement
    "at_4j_arret": 12, "nouvelles_ip": 13, "deces": 16, "journees_it": 17,
}

YEARLY_MP_COLS = {
    "naf5": 2, "nb_salaries": 12, "nb_heures": 13, "nb_siret": 14,
    "events": 15,  # mp_1er_reglement
    "nouvelles_ip": 16, "deces": 19, "journees_it": 20,
}


def parse_yearly_xlsx(xlsx_path, col_map, is_mp=False):
    """Extrait les stats de base par NAF5 depuis un fichier Excel annuel.
    Pour MP, deduplique la main-d'oeuvre par CTN+NAF (meme logique que le pipeline principal).
    Retourne {naf5: {events, nb_salaries, nb_heures, nb_siret, IF, TG}}.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    if is_mp:
        workforce = {}
        naf5_stats = defaultdict(lambda: {
            "events": 0, "nouvelles_ip": 0, "deces": 0, "journees_it": 0,
        })
        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            naf5_val = row[col_map["naf5"]]
            if not naf5_val:
                continue
            naf5 = str(naf5_val).strip()
            if not naf5 or naf5 == "None":
                continue
            ctn = str(row[0] or "").strip()
            key = (ctn, naf5)
            if key not in workforce:
                workforce[key] = {
                    "nb_salaries": safe_num(row[col_map["nb_salaries"]]),
                    "nb_heures": safe_num(row[col_map["nb_heures"]]),
                    "nb_siret": safe_num(row[col_map["nb_siret"]]),
                }
            s = naf5_stats[naf5]
            s["events"] += safe_num(row[col_map["events"]])
            s["nouvelles_ip"] += safe_num(row[col_map["nouvelles_ip"]])
            s["deces"] += safe_num(row[col_map["deces"]])
            s["journees_it"] += safe_num(row[col_map["journees_it"]])

        naf5_wf = defaultdict(lambda: {"nb_salaries": 0, "nb_heures": 0, "nb_siret": 0})
        for (ctn, naf5), wf in workforce.items():
            for k in ["nb_salaries", "nb_heures", "nb_siret"]:
                naf5_wf[naf5][k] += wf[k]

        result = {}
        for naf5, s in naf5_stats.items():
            wf = naf5_wf[naf5]
            nb_sal = wf["nb_salaries"]
            nb_h = wf["nb_heures"]
            result[naf5] = {
                "events": int(s["events"]),
                "nb_salaries": int(nb_sal), "nb_heures": int(nb_h), "nb_siret": int(wf["nb_siret"]),
                "nouvelles_ip": int(s["nouvelles_ip"]), "deces": int(s["deces"]),
                "journees_it": int(s["journees_it"]),
                "indice_frequence": round(s["events"] / nb_sal * 1000, 1) if nb_sal > 0 else 0,
                "taux_gravite": round(s["journees_it"] / (nb_h / 1000), 2) if nb_h > 0 else 0,
            }
    else:
        # AT : agregation simple par NAF5 (somme sur les CTNs)
        naf5_agg = defaultdict(lambda: {
            "events": 0, "at_4j_arret": 0, "nb_salaries": 0, "nb_heures": 0,
            "nb_siret": 0, "nouvelles_ip": 0, "deces": 0, "journees_it": 0,
        })
        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            naf5_val = row[col_map["naf5"]]
            if not naf5_val:
                continue
            naf5 = str(naf5_val).strip()
            if not naf5 or naf5 == "None":
                continue
            a = naf5_agg[naf5]
            for field in ["events", "at_4j_arret", "nb_salaries", "nb_heures",
                          "nb_siret", "nouvelles_ip", "deces", "journees_it"]:
                a[field] += safe_num(row[col_map[field]])

        result = {}
        for naf5, a in naf5_agg.items():
            nb_sal = a["nb_salaries"]
            nb_h = a["nb_heures"]
            result[naf5] = {
                "events": int(a["events"]),
                "nb_salaries": int(nb_sal), "nb_heures": int(nb_h), "nb_siret": int(a["nb_siret"]),
                "nouvelles_ip": int(a["nouvelles_ip"]), "deces": int(a["deces"]),
                "journees_it": int(a["journees_it"]),
                "indice_frequence": round(a["at_4j_arret"] / nb_sal * 1000, 1) if nb_sal > 0 else 0,
                "taux_gravite": round(a["journees_it"] / (nb_h / 1000), 2) if nb_h > 0 else 0,
            }

    wb.close()
    print(f"  [parse] {len(result)} entrees NAF5 depuis {xlsx_path.name}")
    return result


def aggregate_yearly_to_level(yearly_naf5, level_fn):
    """Agrege les donnees annuelles NAF5 au niveau NAF4 ou NAF2."""
    groups = defaultdict(lambda: {
        "events": 0, "nb_salaries": 0, "nb_heures": 0, "nb_siret": 0,
        "nouvelles_ip": 0, "deces": 0, "journees_it": 0,
    })
    for naf5, s in yearly_naf5.items():
        key = level_fn(naf5)
        g = groups[key]
        for field in ["events", "nb_salaries", "nb_heures", "nb_siret",
                       "nouvelles_ip", "deces", "journees_it"]:
            g[field] += s[field]
    result = {}
    for code, g in groups.items():
        nb_sal = g["nb_salaries"]
        nb_h = g["nb_heures"]
        result[code] = {
            **g,
            "indice_frequence": round(g["events"] / nb_sal * 1000, 1) if nb_sal > 0 else 0,
            "taux_gravite": round(g["journees_it"] / (nb_h / 1000), 2) if nb_h > 0 else 0,
        }
    return result


def compute_yearly_national(yearly_naf5):
    """Calcule les totaux nationaux depuis les donnees annuelles NAF5."""
    totals = {"events": 0, "nb_salaries": 0, "nb_heures": 0, "nb_siret": 0,
              "nouvelles_ip": 0, "deces": 0, "journees_it": 0}
    for s in yearly_naf5.values():
        for field in totals:
            totals[field] += s[field]
    nb_sal = totals["nb_salaries"]
    nb_h = totals["nb_heures"]
    totals["indice_frequence"] = round(totals["events"] / nb_sal * 1000, 1) if nb_sal > 0 else 0
    totals["taux_gravite"] = round(totals["journees_it"] / (nb_h / 1000), 2) if nb_h > 0 else 0
    return totals


def merge_yearly_into_data(data, yearly_data_by_year):
    """Ajoute les stats annuelles a chaque entree NAF et aux metadonnees.
    yearly_data_by_year: {"2021": {naf5: {...}, naf4: {...}, naf2: {...}, national: {...}}, "2023": ...}
    """
    years = sorted(yearly_data_by_year.keys())
    data["meta"]["years"] = years
    data["meta"]["national"]["yearly"] = {
        yr: yearly_data_by_year[yr]["national"] for yr in years
    }
    for level in ["by_naf5", "by_naf4", "by_naf2"]:
        level_key = level.replace("by_", "")  # naf5, naf4, naf2
        for code, entry in data[level].items():
            yearly = {}
            for yr in years:
                yr_level = yearly_data_by_year[yr].get(level_key, {})
                if code in yr_level:
                    yearly[yr] = yr_level[code]
            if yearly:
                entry["yearly"] = yearly


# ═══════════════════════════════════════════
# INTEGRATION PDF (optionnel)
# ═══════════════════════════════════════════

AGE_GROUPS = ["<20", "20-24", "25-29", "30-34", "35-39", "40-49", "50-59", "60-64", "65+"]


def merge_pdf_data(at_data, pdf_data, mp_data=None):
    """Fusionne les demographics extraites des PDFs dans les donnees AT (et MP optionnel)."""
    # Niveau NAF5 : fusion directe
    for code, parsed in pdf_data.items():
        if code in at_data["by_naf5"]:
            entry = at_data["by_naf5"][code]
            if parsed["sex"]:
                entry["demographics"] = {"sex": parsed["sex"], "age": parsed["age"]}
            trajet = parsed.get("synthesis", {}).get("trajet")
            if trajet:
                entry["trajet"] = trajet

        if mp_data and code in mp_data["by_naf5"]:
            mp_sex = parsed.get("mp_sex", {})
            mp_age = parsed.get("mp_age", {})
            if mp_sex:
                mp_data["by_naf5"][code]["demographics"] = {"sex": mp_sex, "age": mp_age}

    # Niveaux NAF4/NAF2 : agregation des demographics par somme des NAF5
    for level, key_fn in [("by_naf4", lambda c: c[:4]), ("by_naf2", lambda c: c[:2])]:
        at_agg = defaultdict(lambda: {
            "sex": defaultdict(int), "age": defaultdict(int),
            "trajet_count": 0,
        })
        mp_agg = defaultdict(lambda: {
            "sex": defaultdict(int), "age": defaultdict(int),
        })
        for code, parsed in pdf_data.items():
            key = key_fn(code)
            if parsed["sex"]:
                for s, v in parsed["sex"].items():
                    at_agg[key]["sex"][s] += v
                for a, v in parsed["age"].items():
                    at_agg[key]["age"][a] += v
            trajet = parsed.get("synthesis", {}).get("trajet")
            if trajet:
                at_agg[key]["trajet_count"] += trajet["count"]

            mp_sex = parsed.get("mp_sex", {})
            mp_age = parsed.get("mp_age", {})
            if mp_sex:
                for s, v in mp_sex.items():
                    mp_agg[key]["sex"][s] += v
                for a, v in mp_age.items():
                    mp_agg[key]["age"][a] += v

        for code, data in at_agg.items():
            if code not in at_data[level]:
                continue
            entry = at_data[level][code]
            if data["sex"]:
                entry["demographics"] = {
                    "sex": dict(data["sex"]),
                    "age": dict(data["age"]),
                }
            entry["trajet"] = {"count": data["trajet_count"]}

        if mp_data:
            for code, data in mp_agg.items():
                if code not in mp_data[level]:
                    continue
                if data["sex"]:
                    mp_data[level][code]["demographics"] = {
                        "sex": dict(data["sex"]),
                        "age": dict(data["age"]),
                    }

    # Niveau national : somme de tous les NAF5
    nat_sex = defaultdict(int)
    nat_age = defaultdict(int)
    nat_trajet = 0
    mp_nat_sex = defaultdict(int)
    mp_nat_age = defaultdict(int)
    for parsed in pdf_data.values():
        if parsed["sex"]:
            for s, v in parsed["sex"].items():
                nat_sex[s] += v
            for a, v in parsed["age"].items():
                nat_age[a] += v
        trajet = parsed.get("synthesis", {}).get("trajet")
        if trajet:
            nat_trajet += trajet["count"]

        mp_sex = parsed.get("mp_sex", {})
        mp_age = parsed.get("mp_age", {})
        if mp_sex:
            for s, v in mp_sex.items():
                mp_nat_sex[s] += v
            for a, v in mp_age.items():
                mp_nat_age[a] += v

    at_data["meta"]["national"]["demographics"] = {
        "sex": dict(nat_sex),
        "age": dict(nat_age),
    }
    at_data["meta"]["national"]["trajet"] = {"count": nat_trajet}

    if mp_data:
        mp_data["meta"]["national"]["demographics"] = {
            "sex": dict(mp_nat_sex),
            "age": dict(mp_nat_age),
        }


# ═══════════════════════════════════════════
# PIPELINE TRAJET
# ═══════════════════════════════════════════

def compute_trajet_stats(group):
    """Calcule les statistiques trajet derivees depuis les sommes brutes."""
    nb_sal = group["nb_salaries"]
    count = group["trajet_count"]
    return {
        "trajet_count": int(count),
        "nouvelles_ip": int(group["nouvelles_ip"]),
        "deces": int(group["deces"]),
        "journees_it": int(group["journees_it"]),
        "nb_salaries": int(nb_sal),
        "nb_siret": int(group["nb_siret"]),
        "indice_frequence": round(count / nb_sal * 1000, 1) if nb_sal > 0 else 0,
    }


def build_trajet_data(pdf_data, at_data):
    """Construit les donnees trajet depuis les fiches PDF + main-d'oeuvre AT.

    Utilise les valeurs 2023 de trajet_yearly pour les stats ; 5 ans pour l'evolution.
    Main-d'oeuvre (nb_salaries, nb_siret) issue des donnees AT.
    """
    years = ["2019", "2020", "2021", "2022", "2023"]

    # Niveau NAF5
    by_naf5 = {}
    for code, pdf in pdf_data.items():
        yearly = pdf.get("trajet_yearly")
        if not yearly:
            continue
        at_entry = at_data["by_naf5"].get(code)
        if not at_entry:
            continue
        at_stats = at_entry["stats"]
        y23 = yearly["2023"]

        entry = {
            "libelle": at_entry["libelle"],
            "naf4": code[:4],
            "naf2": code[:2],
            "stats": compute_trajet_stats({
                "trajet_count": y23["count"],
                "nouvelles_ip": y23["ip"],
                "deces": y23["deces"],
                "journees_it": y23["journees"],
                "nb_salaries": at_stats["nb_salaries"],
                "nb_siret": at_stats["nb_siret"],
            }),
            "yearly": {},
        }
        for yr in years:
            if yr in yearly:
                y = yearly[yr]
                at_yr = at_entry.get("yearly", {}).get(yr)
                nb_sal = at_yr["nb_salaries"] if at_yr else at_stats["nb_salaries"]
                entry["yearly"][yr] = {
                    "events": y["count"],
                    "nb_salaries": int(nb_sal),
                    "nouvelles_ip": y["ip"],
                    "deces": y["deces"],
                    "journees_it": y["journees"],
                    "indice_frequence": round(y["count"] / nb_sal * 1000, 1) if nb_sal > 0 else 0,
                }
        by_naf5[code] = entry

    # Agregation vers NAF4
    naf4_agg = defaultdict(lambda: {
        "trajet_count": 0, "nouvelles_ip": 0, "deces": 0, "journees_it": 0,
        "nb_salaries": 0, "nb_siret": 0, "libelle": "", "source_codes": [],
        "yearly_agg": defaultdict(lambda: {
            "events": 0, "nb_salaries": 0, "nouvelles_ip": 0, "deces": 0, "journees_it": 0,
        }),
    })
    for code, entry in by_naf5.items():
        key = code[:4]
        g = naf4_agg[key]
        s = entry["stats"]
        g["trajet_count"] += s["trajet_count"]
        g["nouvelles_ip"] += s["nouvelles_ip"]
        g["deces"] += s["deces"]
        g["journees_it"] += s["journees_it"]
        g["nb_salaries"] += s["nb_salaries"]
        g["nb_siret"] += s["nb_siret"]
        g["source_codes"].append(code)
        if not g["libelle"]:
            g["libelle"] = entry["libelle"]
        for yr, yd in entry.get("yearly", {}).items():
            ya = g["yearly_agg"][yr]
            ya["events"] += yd["events"]
            ya["nb_salaries"] += yd["nb_salaries"]
            ya["nouvelles_ip"] += yd["nouvelles_ip"]
            ya["deces"] += yd["deces"]
            ya["journees_it"] += yd["journees_it"]

    by_naf4 = {}
    for code, g in sorted(naf4_agg.items()):
        yearly = {}
        for yr, ya in g["yearly_agg"].items():
            nb_sal = ya["nb_salaries"]
            yearly[yr] = {
                **ya,
                "indice_frequence": round(ya["events"] / nb_sal * 1000, 1) if nb_sal > 0 else 0,
            }
        by_naf4[code] = {
            "libelle": g["libelle"], "naf2": code[:2],
            "codes_naf5": sorted(set(g["source_codes"])),
            "stats": compute_trajet_stats(g),
            "yearly": yearly,
        }

    # Agregation vers NAF2
    naf2_agg = defaultdict(lambda: {
        "trajet_count": 0, "nouvelles_ip": 0, "deces": 0, "journees_it": 0,
        "nb_salaries": 0, "nb_siret": 0, "libelle": "",
        "yearly_agg": defaultdict(lambda: {
            "events": 0, "nb_salaries": 0, "nouvelles_ip": 0, "deces": 0, "journees_it": 0,
        }),
    })
    for code, entry in by_naf5.items():
        key = code[:2]
        g = naf2_agg[key]
        s = entry["stats"]
        g["trajet_count"] += s["trajet_count"]
        g["nouvelles_ip"] += s["nouvelles_ip"]
        g["deces"] += s["deces"]
        g["journees_it"] += s["journees_it"]
        g["nb_salaries"] += s["nb_salaries"]
        g["nb_siret"] += s["nb_siret"]
        if not g["libelle"]:
            at_naf2 = at_data["by_naf2"].get(key)
            g["libelle"] = at_naf2["libelle"] if at_naf2 else entry["libelle"]
        for yr, yd in entry.get("yearly", {}).items():
            ya = g["yearly_agg"][yr]
            ya["events"] += yd["events"]
            ya["nb_salaries"] += yd["nb_salaries"]
            ya["nouvelles_ip"] += yd["nouvelles_ip"]
            ya["deces"] += yd["deces"]
            ya["journees_it"] += yd["journees_it"]

    by_naf2 = {}
    for code, g in sorted(naf2_agg.items()):
        yearly = {}
        for yr, ya in g["yearly_agg"].items():
            nb_sal = ya["nb_salaries"]
            yearly[yr] = {
                **ya,
                "indice_frequence": round(ya["events"] / nb_sal * 1000, 1) if nb_sal > 0 else 0,
            }
        by_naf2[code] = {
            "libelle": g["libelle"],
            "stats": compute_trajet_stats(g),
            "yearly": yearly,
        }

    # National
    national = {
        "trajet_count": 0, "nouvelles_ip": 0, "deces": 0, "journees_it": 0,
        "nb_salaries": 0, "nb_siret": 0,
    }
    nat_yearly_agg = defaultdict(lambda: {
        "events": 0, "nb_salaries": 0, "nouvelles_ip": 0, "deces": 0, "journees_it": 0,
    })
    for entry in by_naf5.values():
        s = entry["stats"]
        national["trajet_count"] += s["trajet_count"]
        national["nouvelles_ip"] += s["nouvelles_ip"]
        national["deces"] += s["deces"]
        national["journees_it"] += s["journees_it"]
        national["nb_salaries"] += s["nb_salaries"]
        national["nb_siret"] += s["nb_siret"]
        for yr, yd in entry.get("yearly", {}).items():
            ya = nat_yearly_agg[yr]
            ya["events"] += yd["events"]
            ya["nb_salaries"] += yd["nb_salaries"]
            ya["nouvelles_ip"] += yd["nouvelles_ip"]
            ya["deces"] += yd["deces"]
            ya["journees_it"] += yd["journees_it"]

    national_stats = compute_trajet_stats(national)
    nat_yearly = {}
    for yr, ya in sorted(nat_yearly_agg.items()):
        nb_sal = ya["nb_salaries"]
        nat_yearly[yr] = {
            **ya,
            "indice_frequence": round(ya["events"] / nb_sal * 1000, 1) if nb_sal > 0 else 0,
        }
    national_stats["yearly"] = nat_yearly

    naf_index = []
    for code, data in by_naf5.items():
        naf_index.append({"code": code, "libelle": data["libelle"], "level": "naf5"})
    for code, data in by_naf4.items():
        naf_index.append({"code": code, "libelle": data["libelle"], "level": "naf4"})
    for code, data in by_naf2.items():
        naf_index.append({"code": code, "libelle": data["libelle"], "level": "naf2"})
    naf_index.sort(key=lambda x: (x["level"], x["code"]))

    return {
        "meta": {
            "source": "Ameli, Fiches NAF 2023 (PDF) + main-d'oeuvre AT",
            "source_url": "https://assurance-maladie.ameli.fr/etudes-et-donnees/sinistralite-at-mp-par-code-naf",
            "years": years,
            "national": national_stats,
        },
        "by_naf5": by_naf5, "by_naf4": by_naf4, "by_naf2": by_naf2,
        "naf_index": naf_index,
    }


def build_yearly_from_pdf(pdf_data, section_key, base_data=None):
    """Construit l'evolution annuelle depuis les tableaux PDF (5 ans).

    Args:
        pdf_data: {naf5: parsed_pdf} depuis parse_all_pdfs
        section_key: "at_yearly" ou "mp_yearly"
        base_data: pour MP, passer les donnees AT pour obtenir la main-d'oeuvre annuelle.
                   AT utilise ses propres salaries issus du PDF ; fallback sur l'Excel 2023.

    Retourne: {year: {naf5: {...}, naf4: {...}, naf2: {...}, national: {...}}, ...}
    """
    years = ["2019", "2020", "2021", "2022", "2023"]
    yearly_by_year = {}

    for year in years:
        naf5_data = {}
        for code, pdf in pdf_data.items():
            yearly = pdf.get(section_key)
            if not yearly or year not in yearly:
                continue
            y = yearly[year]

            if section_key == "at_yearly":
                nb_sal = y.get("salaries", 0)
                if nb_sal == 0 and base_data and code in base_data["by_naf5"]:
                    nb_sal = base_data["by_naf5"][code]["stats"]["nb_salaries"]
            else:
                nb_sal = 0
                at_yearly = pdf_data.get(code, {}).get("at_yearly")
                if at_yearly and year in at_yearly:
                    nb_sal = at_yearly[year].get("salaries", 0)
                if nb_sal == 0 and base_data and code in base_data["by_naf5"]:
                    nb_sal = base_data["by_naf5"][code]["stats"]["nb_salaries"]

            nb_h = 0
            if base_data and code in base_data["by_naf5"]:
                nb_h = base_data["by_naf5"][code]["stats"].get("nb_heures", 0)

            events = y["count"]
            journees = y["journees"]
            naf5_data[code] = {
                "events": events,
                "nb_salaries": int(nb_sal),
                "nb_heures": int(nb_h),
                "nb_siret": 0,
                "nouvelles_ip": y["ip"],
                "deces": y["deces"],
                "journees_it": journees,
                "indice_frequence": round(events / nb_sal * 1000, 1) if nb_sal > 0 else 0,
                "taux_gravite": round(journees / (nb_h / 1000), 2) if nb_h > 0 else 0,
            }

        yearly_by_year[year] = {
            "naf5": naf5_data,
            "naf4": aggregate_yearly_to_level(naf5_data, lambda c: c[:4]),
            "naf2": aggregate_yearly_to_level(naf5_data, lambda c: c[:2]),
            "national": compute_yearly_national(naf5_data),
        }

    count = sum(1 for pdf in pdf_data.values() if pdf.get(section_key))
    print(f"  [build] Evolution 5 ans depuis {count} PDFs ({section_key})")
    return yearly_by_year


def validate(data, label, spot_code="4711D"):
    """Affiche un resume de validation."""
    print(f"\n  [valid] {label} :")
    print(f"    NAF5 : {len(data['by_naf5'])} | NAF4 : {len(data['by_naf4'])} | NAF2 : {len(data['by_naf2'])}")
    print(f"    Index : {len(data['naf_index'])} entrees")
    nat = data["meta"]["national"]
    tg = nat.get("taux_gravite", "n/a")
    print(f"    National IF : {nat['indice_frequence']} | TG : {tg}")

    if spot_code in data["by_naf5"]:
        d = data["by_naf5"][spot_code]
        s = d["stats"]
        for ek in ["at_1er_reglement", "mp_1er_reglement", "trajet_count"]:
            if ek in s:
                event_key = ek
                break
        print(f"    Spot {spot_code} ({d['libelle']}) :")
        print(f"      Salaries : {s['nb_salaries']:,} | {event_key} : {s[event_key]:,}")
        print(f"      IF : {s['indice_frequence']} | IP : {s['nouvelles_ip']}")


# ═══════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Rafraichit les donnees AT/MP/Trajet depuis ameli.fr."
    )
    parser.add_argument(
        "--pdf-dir",
        help="Dossier contenant les PDF NAF_*.pdf (optionnel, pour les demographics et le Trajet).",
        default=None,
    )
    args = parser.parse_args()

    pdf_dir = None
    if args.pdf_dir:
        pdf_dir = Path(args.pdf_dir)
        if not pdf_dir.exists():
            print(f"[erreur] Dossier PDF introuvable : {pdf_dir}", file=sys.stderr)
            sys.exit(1)

    # ── AT ──
    print("=== Pipeline AT ===")
    print("[1] Telechargement...")
    download_xlsx(AT_XLSX_PATH, AT_XLSX_URL)
    # Le fichier 2021 fournit les colonnes de causes de risque (col. 20-31)
    # absentes du format 2023. Necessaire pour les graphiques de causes AT.
    download_xlsx(AT_2021_XLSX_PATH, AT_2021_XLSX_URL)
    print("[2] Parsing...")
    at_rows = parse_at_xlsx()
    print("[3] Construction...")
    at_data = build_at_data(at_rows)

    # ── MP ──
    print("\n=== Pipeline MP ===")
    print("[1] Telechargement...")
    download_xlsx(MP_XLSX_PATH, MP_XLSX_URL)
    print("[2] Parsing...")
    mp_rows = parse_mp_xlsx()
    print("[3] Construction...")
    mp_data = build_mp_data(mp_rows)

    # ── PDF (optionnel) ──
    if pdf_dir is not None:
        try:
            from parse_pdf import parse_all_pdfs
        except ImportError:
            print("[avert] parse_pdf.py introuvable - demographics et Trajet ignores.")
            print("        Verifiez que parse_pdf.py est present dans data/pipeline/.")
            pdf_dir = None

    if pdf_dir is not None:
        print("\n=== Parsing PDF ===")
        pdf_data = parse_all_pdfs(pdf_dir)
        merge_pdf_data(at_data, pdf_data, mp_data=mp_data)

        print("\n=== Evolution AT (5 ans) ===")
        at_yearly = build_yearly_from_pdf(pdf_data, "at_yearly", base_data=at_data)
        merge_yearly_into_data(at_data, at_yearly)

        print("\n=== Evolution MP (5 ans) ===")
        mp_yearly = build_yearly_from_pdf(pdf_data, "mp_yearly", base_data=at_data)
        merge_yearly_into_data(mp_data, mp_yearly)
    else:
        if args.pdf_dir is None:
            print("\n[info] --pdf-dir non fourni. Demographics et Trajet non generes.")
            print("       Pour les inclure : python refresh_data.py --pdf-dir /chemin/vers/pdfs")

    # ── Ecriture AT ──
    print("\n[write] AT...")
    write_json(at_data, AT_JSON_PATH, "AT")
    validate(at_data, "AT")

    # ── Ecriture MP ──
    print("[write] MP...")
    write_json(mp_data, MP_JSON_PATH, "MP")
    validate(mp_data, "MP")

    # ── Trajet (necessite les PDFs) ──
    if pdf_dir is not None:
        print("\n=== Pipeline Trajet ===")
        trajet_data = build_trajet_data(pdf_data, at_data)
        write_json(trajet_data, TRAJET_JSON_PATH, "Trajet")
        validate(trajet_data, "Trajet")
    else:
        print("\n[info] Trajet non genere (necessite --pdf-dir).")

    print("\nTermine.")


if __name__ == "__main__":
    main()
